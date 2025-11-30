"""
Exporter - Modul zum Export von Aufmaß-Listen in verschiedene Formate.

Dieses Modul exportiert die berechneten Mengen als:
- Excel (.xlsx)
- CSV
- JSON
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)


class Exporter:
    """
    Klasse zum Export von Aufmaß-Listen.
    
    Unterstützt folgende Formate:
    - Excel (.xlsx) mit formatierter Tabelle
    - CSV für einfachen Datenaustausch
    - JSON für maschinenlesbare Ausgabe
    
    Attributes:
        project_name: Name des Projekts für Header
        created_by: Ersteller-Name
    """
    
    def __init__(
        self,
        project_name: str = "Bauzeichnung-Analyse",
        created_by: str = "BauzeichnungAnalyzer"
    ) -> None:
        """
        Initialisiert den Exporter.
        
        Args:
            project_name: Name des Projekts
            created_by: Ersteller-Name
        """
        self.project_name = project_name
        self.created_by = created_by
        logger.info(f"Exporter initialisiert für Projekt: {project_name}")
    
    def export_to_excel(
        self,
        quantity_result: Any,
        output_path: Union[str, Path],
        include_summary: bool = True
    ) -> Path:
        """
        Exportiert Aufmaß-Liste als Excel-Datei.
        
        Args:
            quantity_result: QuantityResult-Objekt
            output_path: Ausgabepfad für die Excel-Datei
            include_summary: Ob eine Zusammenfassung hinzugefügt werden soll
            
        Returns:
            Pfad zur erstellten Excel-Datei
        """
        output_path = Path(output_path)
        
        if not output_path.suffix:
            output_path = output_path.with_suffix(".xlsx")
        
        logger.info(f"Exportiere als Excel: {output_path}")
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Erstelle Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Aufmaß-Liste"
            
            # Styles definieren
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Header
            ws.merge_cells('A1:F1')
            ws['A1'] = f"Aufmaß-Liste: {self.project_name}"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A2'] = f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A3'] = f"Erstellt von: {self.created_by}"
            
            # Spaltenüberschriften
            headers = ["Pos.", "Kategorie", "Beschreibung", "Menge", "Einheit", "Anmerkungen"]
            row_num = 5
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Daten schreiben
            for item in quantity_result.items:
                row_num += 1
                ws.cell(row=row_num, column=1, value=item.position).border = thin_border
                ws.cell(row=row_num, column=2, value=item.category).border = thin_border
                ws.cell(row=row_num, column=3, value=item.description).border = thin_border
                ws.cell(row=row_num, column=4, value=round(item.quantity, 2)).border = thin_border
                ws.cell(row=row_num, column=5, value=item.unit).border = thin_border
                ws.cell(row=row_num, column=6, value=item.notes).border = thin_border
            
            # Zusammenfassung
            if include_summary and quantity_result.summary:
                row_num += 3
                ws.cell(row=row_num, column=1, value="Zusammenfassung").font = Font(bold=True, size=12)
                row_num += 1
                
                for category, units in quantity_result.summary.items():
                    for unit, total in units.items():
                        row_num += 1
                        ws.cell(row=row_num, column=2, value=category)
                        ws.cell(row=row_num, column=4, value=round(total, 2))
                        ws.cell(row=row_num, column=5, value=unit)
            
            # Spaltenbreiten anpassen
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 40
            
            # Speichern
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path))
            
            logger.info(f"Excel-Export erfolgreich: {output_path}")
            return output_path
            
        except ImportError as e:
            logger.error(f"Erforderliche Bibliothek nicht installiert: {e}")
            raise
    
    def export_to_csv(
        self,
        quantity_result: Any,
        output_path: Union[str, Path],
        delimiter: str = ";"
    ) -> Path:
        """
        Exportiert Aufmaß-Liste als CSV-Datei.
        
        Args:
            quantity_result: QuantityResult-Objekt
            output_path: Ausgabepfad für die CSV-Datei
            delimiter: Trennzeichen (Standard: ";")
            
        Returns:
            Pfad zur erstellten CSV-Datei
        """
        output_path = Path(output_path)
        
        if not output_path.suffix:
            output_path = output_path.with_suffix(".csv")
        
        logger.info(f"Exportiere als CSV: {output_path}")
        
        try:
            import pandas as pd
            
            # DataFrame erstellen
            data = []
            for item in quantity_result.items:
                data.append({
                    "Position": item.position,
                    "Kategorie": item.category,
                    "Beschreibung": item.description,
                    "Menge": round(item.quantity, 2),
                    "Einheit": item.unit,
                    "Anmerkungen": item.notes
                })
            
            df = pd.DataFrame(data)
            
            # Als CSV speichern
            output_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(str(output_path), sep=delimiter, index=False, encoding='utf-8-sig')
            
            logger.info(f"CSV-Export erfolgreich: {output_path}")
            return output_path
            
        except ImportError:
            # Fallback ohne pandas
            logger.warning("pandas nicht installiert, verwende Standard-CSV")
            
            import csv
            
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=delimiter)
                
                # Header
                writer.writerow(["Position", "Kategorie", "Beschreibung", "Menge", "Einheit", "Anmerkungen"])
                
                # Daten
                for item in quantity_result.items:
                    writer.writerow([
                        item.position,
                        item.category,
                        item.description,
                        round(item.quantity, 2),
                        item.unit,
                        item.notes
                    ])
            
            logger.info(f"CSV-Export erfolgreich: {output_path}")
            return output_path
    
    def export_to_json(
        self,
        quantity_result: Any,
        output_path: Union[str, Path],
        indent: int = 2
    ) -> Path:
        """
        Exportiert Aufmaß-Liste als JSON-Datei.
        
        Args:
            quantity_result: QuantityResult-Objekt
            output_path: Ausgabepfad für die JSON-Datei
            indent: Einrückung für JSON-Formatierung
            
        Returns:
            Pfad zur erstellten JSON-Datei
        """
        output_path = Path(output_path)
        
        if not output_path.suffix:
            output_path = output_path.with_suffix(".json")
        
        logger.info(f"Exportiere als JSON: {output_path}")
        
        # Daten vorbereiten
        export_data = {
            "meta": {
                "project_name": self.project_name,
                "created_by": self.created_by,
                "created_at": datetime.now().isoformat(),
                "total_items": len(quantity_result.items)
            },
            "items": [],
            "summary": quantity_result.summary
        }
        
        for item in quantity_result.items:
            export_data["items"].append({
                "position": item.position,
                "category": item.category,
                "description": item.description,
                "quantity": round(item.quantity, 2),
                "unit": item.unit,
                "dimensions": {k: round(v, 4) for k, v in item.dimensions.items()},
                "notes": item.notes
            })
        
        # Als JSON speichern
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=indent, ensure_ascii=False)
        
        logger.info(f"JSON-Export erfolgreich: {output_path}")
        return output_path
    
    def export_all(
        self,
        quantity_result: Any,
        output_dir: Union[str, Path],
        base_name: str = "aufmass"
    ) -> Dict[str, Path]:
        """
        Exportiert Aufmaß-Liste in allen verfügbaren Formaten.
        
        Args:
            quantity_result: QuantityResult-Objekt
            output_dir: Ausgabeverzeichnis
            base_name: Basis-Dateiname (ohne Erweiterung)
            
        Returns:
            Dictionary mit Format als Schlüssel und Pfad als Wert
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"Exportiere in alle Formate nach: {output_dir}")
        
        results = {}
        
        # Excel
        try:
            excel_path = self.export_to_excel(
                quantity_result,
                output_dir / f"{base_name}.xlsx"
            )
            results["excel"] = excel_path
        except Exception as e:
            logger.error(f"Excel-Export fehlgeschlagen: {e}")
        
        # CSV
        try:
            csv_path = self.export_to_csv(
                quantity_result,
                output_dir / f"{base_name}.csv"
            )
            results["csv"] = csv_path
        except Exception as e:
            logger.error(f"CSV-Export fehlgeschlagen: {e}")
        
        # JSON
        try:
            json_path = self.export_to_json(
                quantity_result,
                output_dir / f"{base_name}.json"
            )
            results["json"] = json_path
        except Exception as e:
            logger.error(f"JSON-Export fehlgeschlagen: {e}")
        
        logger.info(f"Export abgeschlossen: {len(results)} Format(e) erstellt")
        return results
